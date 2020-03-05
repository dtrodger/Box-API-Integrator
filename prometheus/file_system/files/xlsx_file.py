from __future__ import annotations
import logging
from typing import AnyStr, Optional, List
import os
import datetime
import uuid

import openpyxl
import boxsdk
import attr

from prometheus.file_system.files import prometheus_file


log = logging.getLogger(__name__)


@attr.s(slots=True, auto_attribs=True)
class XLSXFile(prometheus_file.PrometheusFile):
    file_path: str = None
    attempted_upload_to_box_on: Optional[datetime.datetime] = None
    box_file: Optional[boxsdk.object.file.File] = None
    box_file_id: Optional[str] = None
    box_upload_folder_id: Optional[str] = None
    box_upload_folder_path: Optional[str] = None
    box_upload_user: Optional[boxsdk.object.user.User] = None
    box_upload_user_email: Optional[str] = None
    box_upload_user_id: Optional[str] = None
    current_directory: Optional[str] = None
    deleted: bool = False
    deleted_on: bool = False
    directory_entry_on: Optional[datetime.datetime] = None
    file_directory_path: Optional[str] = None
    file_name: Optional[str] = None
    id: uuid.UUID = uuid.uuid4()
    initial_monitor_on: Optional[datetime.datetime] = None
    input_complete: bool = False
    min_elapsed_for_box_upload_attempt: int = 1
    min_elapsed_for_box_upload_fail: int = 3
    min_elapsed_for_delete: int = 0
    min_elapsed_for_input_complete: int = 1
    ready_for_delete: bool = False
    ready_for_upload: bool = False
    st_check_on: Optional[datetime.datetime] = None
    st_creation_time: Optional[int] = None
    st_last_accessed_time: Optional[int] = None
    st_last_modified_time: Optional[int] = None
    st_size: Optional[int] = None
    st_size_diff_from_cache_on: Optional[datetime.datetime] = None
    xlsx_workbook: Optional[openpyxl.reader.excel.ExcelReader] = None
    uploaded_to_box: bool = False
    uploaded_to_box_on: datetime = None

    def append_row(self, row_fields: List, worksheet_name: AnyStr = "Sheet") -> None:
        log.debug(f"{self} appending row {row_fields}")
        self.xlsx_workbook[worksheet_name].append(row_fields)
        self.save_workbook()

    def create_on_disk(self, override_existing: bool = True) -> bool:
        created = False
        if override_existing or not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            wb.save(self.file_path)
            created = True

        if created:
            log.debug(f"created file {str(self)} on disk")
        else:
            log.debug(f"did not create file {str(self)} on disk")

        return created

    def next_empty_row_index(self, worksheet_name: AnyStr = "Sheet") -> int:
        return self.xlsx_workbook[worksheet_name].max_row + 1

    def save_workbook(self, file_path: AnyStr = None) -> bool:
        saved = False
        try:
            if file_path:
                self.xlsx_workbook.save(filename=file_path)
            else:
                self.xlsx_workbook.save(filename=self.file_path)
            saved = True
        except Exception as e:
            error = e

        if saved:
            log.debug(f"{str(self)} save xlsx workbook")
        else:
            log.error(f"{str(self)} failed to save xlsx workbook with {str(error)}")

        return saved

    def set_xlsx_workbook(self, file_path: AnyStr = None) -> Optional[openpyxl.reader.excel.ExcelReader]:
        if not file_path:
            file_path = self.file_path

        xlsx_workbook = None
        if os.path.exists(file_path):
            self.xlsx_workbook = openpyxl.load_workbook(file_path)
            xlsx_workbook = self.xlsx_workbook

        return xlsx_workbook
